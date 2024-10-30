import logging
import os
import numpy as np

from openpyxl import load_workbook
from transformers import pipeline
from fastapi import UploadFile
from pathlib import Path
from pdfminer.high_level import extract_text
from sklearn.feature_extraction.text import TfidfVectorizer
from transformers import AutoTokenizer
from api.core.util import Util

# 로거 설정
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class Preprocessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.categories = []
        self.descriptions = []
        self._load_data()

    def _load_data(self):
        # 파일에서 데이터를 로드하고 categories와 descriptions를 채웁니다.
        # 예시로 엑셀 파일에서 데이터를 로드한다고 가정합니다.
        workbook = load_workbook(self.file_path)
        sheet = workbook["context"]

        for row in sheet.iter_rows(values_only=True):
            if row[0] and row[1]:
                self.categories.append(row[0])  # 카테고리
                self.descriptions.append(row[1])  # 설명

    def get_categories(self):
        return self.categories

    def get_descriptions(self):
        return self.descriptions